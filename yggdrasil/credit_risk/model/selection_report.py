"""
yggdrasil.credit_risk.model.selection_report
============================================
**Relatório da seleção de variáveis** — a camada de apresentação do
:class:`~yggdrasil.credit_risk.model.selection.SelectionResult` devolvido pela
:func:`~yggdrasil.credit_risk.model.selection.run_selection`.

A esteira responde *o que* entrou, *onde* cada variável saiu e *por quê*; aqui
isso vira material de **apresentação corporativa**: gráficos, tabelas formatadas
em pt-BR, uma página HTML autocontida (anexável a um comitê) e um Excel
multi-abas para quem quiser continuar a análise na planilha::

    from yggdrasil.credit_risk.model import run_selection
    from yggdrasil.credit_risk.model import (plot_funil, plot_iv_ranking,
                                             save_selection_report)

    res = run_selection(seg, apply=True)
    plot_funil(res)                       # o gráfico-síntese da apresentação
    plot_iv_ranking(res, top=20)          # quem tem sinal, colorido pela decisão
    save_selection_report(res, "selecao.html", seg=seg)

Quatro gráficos, pensados como uma narrativa:

1. :func:`plot_funil` — de quantas candidatas partimos e quantas cada etapa
   levou (a foto de abertura);
2. :func:`plot_motivos` — **por que** perdemos variáveis (a leitura executiva do
   funil);
3. :func:`plot_iv_ranking` — quem tem poder discriminante, com o corte usado;
4. :func:`plot_iv_psi` — poder × estabilidade em quadrantes (a matriz de decisão).

Nada aqui é específico de um parâmetro de risco: o alvo é sempre nomeado pelo
``problema`` da política (o ``problem_label`` do segmentador).

Sobre o HTML — por que uma página própria
-----------------------------------------
:func:`yggdrasil.credit_risk._mlflow_report.build_tabbed_report_html` monta abas
a partir de **DataFrames** e só deles: não há como injetar figuras (data URI) nem
os cartões do sumário executivo, e a paleta dela é fixa em tema claro. Como esse
módulo compartilhado não pode ser alterado por aqui, o relatório escreve o
próprio HTML — mas **reaproveita a linguagem visual** dela (mesma família
tipográfica, mesmo azul de acento, mesmo desenho de tabela), agora com as cores
declaradas como variáveis CSS no ``:root`` e um bloco
``@media (prefers-color-scheme: dark)`` para tema escuro. O arquivo é
**autocontido**: CSS embutido, figuras em ``data:image/png;base64`` e nenhuma
requisição externa.
"""
from __future__ import annotations

import base64
import html as _html
import io
from datetime import datetime

import numpy as np
import pandas as pd

from .selection import COLUNAS_FUNIL, COLUNAS_TABELA, rotulo_etapa

# ----------------------------------------------------------------------
# Vocabulário de apresentação
# ----------------------------------------------------------------------
#: Decisões na ordem em que aparecem no relatório (da melhor para a pior).
ORDEM_DECISAO = ("selecionada", "revisar", "excluida")

#: Rótulo em pt-BR de cada decisão.
DECISAO_LABEL = {"selecionada": "Selecionada", "revisar": "A revisar",
                 "excluida": "Excluída"}

#: Cor de cada decisão nos gráficos — verde/âmbar/vermelho da paleta do projeto
#: (mesmos tons dos plots do segmentador).
DECISAO_COR = {"selecionada": "#157a52", "revisar": "#9a6f12",
               "excluida": "#b23a2a"}

_TIPO_LABEL = {"num": "numérica", "cat": "categórica"}

_AZUL = "#15324a"          # títulos e barra das candidatas
_AZUL_CLARO = "#7ba4c1"    # fim do degradê das etapas
_CINZA = "#889"            # avisos amigáveis no eixo

#: Rótulos legíveis das colunas de ``SelectionResult.tabela``.
ROTULOS_TABELA = {
    "variavel": "Variável", "rotulo": "Rótulo", "tipo": "Tipo",
    "decisao": "Decisão", "etapa_saida": "Etapa", "motivo": "Motivo",
    "iv": "IV", "forca": "Força", "pior_psi": "PSI (pior)",
    "estabilidade": "Estabilidade", "tendencia": "Tendência",
    "n_inversoes": "Inversões", "missing_pct": "Faltantes",
    "n_categorias": "Categorias",
}

#: Rótulos legíveis das colunas de ``SelectionResult.funil``.
ROTULOS_FUNIL = {
    "etapa": "Etapa", "n_entrada": "Entraram", "n_excluidas": "Excluídas",
    "n_revisar": "A revisar", "n_saida": "Seguiram", "retido_pct": "% das candidatas",
}

#: Causa **curta** de exclusão por etapa — a leitura executiva do motivo por
#: extenso (que traz os números do caso). Etapa fora do mapa cai no rótulo dela.
CAUSAS_CURTAS = {
    "missing": "faltantes acima do limite",
    "constante": "sem variabilidade (valor único ou categoria dominante)",
    "categoricas": "cardinalidade alta ou categorias sem massa",
    "iv": "IV abaixo do mínimo",
    "psi": "PSI acima do máximo (instável)",
    "monotonia": "ordem de risco não-monotônica",
    "correlacao": "redundante com outra variável",
    "vif": "multicolinearidade (VIF alto)",
    "backward": "removida no backward elimination",
}


# ======================================================================
# Helpers de formatação (pt-BR) e de figura
# ======================================================================
def _f(x):
    """``float`` finito ou ``None`` (trata ``NaN``/``pd.NA``/texto)."""
    try:
        v = float(x)
    except (TypeError, ValueError):
        return None
    return v if np.isfinite(v) else None


def _fmt_num(x, nd: int = 3) -> str:
    """Número com **vírgula decimal** (pt-BR); ``"—"`` quando ausente."""
    v = _f(x)
    return "—" if v is None else f"{v:.{nd}f}".replace(".", ",")


def _fmt_pct(x, nd: int = 1) -> str:
    """Percentual já em escala 0–100: ``12.34`` → ``"12,3%"``."""
    v = _f(x)
    return "—" if v is None else f"{v:.{nd}f}".replace(".", ",") + "%"


def _fmt_int(x) -> str:
    """Inteiro legível; ``"—"`` quando ausente (``pd.NA``/``NaN``)."""
    v = _f(x)
    return "—" if v is None else f"{int(round(v))}"


def _curto(texto, n: int = 34) -> str:
    """Trunca um rótulo longo para caber no eixo/legenda."""
    s = str(texto)
    return s if len(s) <= n else s[: n - 1] + "…"


def _new_ax(figsize, dpi, ax):
    """Figura SEM pyplot (não entra no Gcf) — evita o backend inline re-exibir.

    Espelha, de propósito, o helper homônimo do segmentador: o relatório é uma
    camada de apresentação e não deve depender de detalhe privado dele."""
    if ax is not None:
        return ax.figure, ax
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    from matplotlib.figure import Figure
    fig = Figure(figsize=figsize, dpi=dpi)
    FigureCanvasAgg(fig)
    return fig, fig.subplots()


def _aviso_no_eixo(fig, ax, msg: str, save_path=None, dpi: int = 150):
    """Mensagem amigável no lugar do gráfico (informação ausente não é erro)."""
    ax.text(0.5, 0.5, msg, ha="center", va="center", transform=ax.transAxes,
            color=_CINZA, fontsize=10, wrap=True)
    ax.axis("off")
    fig.tight_layout()
    if save_path:
        fig.savefig(save_path, dpi=dpi, bbox_inches="tight")
    return fig


def _eixo_inteiro(ax, eixo: str = "x"):
    """Ticks inteiros — contagens de variáveis não têm meia unidade."""
    from matplotlib.ticker import MaxNLocator
    alvo = ax.xaxis if eixo == "x" else ax.yaxis
    alvo.set_major_locator(MaxNLocator(integer=True))


def _finish(fig, save_path=None, dpi: int = 150):
    fig.tight_layout()
    if save_path:
        fig.savefig(save_path, dpi=dpi, bbox_inches="tight")
    return fig


def _degrade(n: int, inicio=_AZUL, fim=_AZUL_CLARO) -> list:
    """Degradê linear entre duas cores hex — ``n`` cores (sem colormap)."""
    def _rgb(h):
        h = h.lstrip("#")
        return tuple(int(h[i:i + 2], 16) / 255.0 for i in (0, 2, 4))
    a, b = _rgb(inicio), _rgb(fim)
    if n <= 1:
        return [a]
    return [tuple(a[k] + (b[k] - a[k]) * i / (n - 1) for k in range(3))
            for i in range(n)]


def fig_to_data_uri(fig, dpi: int = 110) -> str:
    """Figura matplotlib → ``data:image/png;base64,...`` (para o HTML autocontido).

    Usa PNG em ``dpi`` moderado: o relatório costuma viajar por e-mail junto da
    apresentação, e ``dpi`` alto multiplica o tamanho do arquivo sem ganho
    perceptível na tela."""
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    return f"data:image/png;base64,{b64}"


# ----------------------------------------------------------------------
# Leitura defensiva do resultado (funciona também com from_dict)
# ----------------------------------------------------------------------
def _politica(result) -> dict:
    return dict(getattr(result, "politica", None) or {})


def _params(result) -> dict:
    return dict(_politica(result).get("parametros") or {})


def _problema(result) -> str:
    """Rótulo NEUTRO do alvo (``problem_label`` do segmentador)."""
    pol = _politica(result)
    return str(pol.get("problema") or pol.get("alvo") or "alvo")


def _tabela(result) -> pd.DataFrame:
    tab = getattr(result, "tabela", None)
    if tab is None or not len(tab):
        return pd.DataFrame(columns=list(COLUNAS_TABELA))
    return tab.copy()


def _funil(result) -> pd.DataFrame:
    fun = getattr(result, "funil", None)
    if fun is None or not len(fun):
        return pd.DataFrame(columns=list(COLUNAS_FUNIL))
    return fun.copy()


def _por_decisao(result) -> dict:
    """Contagem por decisão, lida da tabela (robusta ao round-trip JSON)."""
    tab = _tabela(result)
    dec = tab["decisao"].astype(str) if len(tab) else pd.Series(dtype=object)
    return {d: int((dec == d).sum()) for d in ORDEM_DECISAO}


def causas_exclusao(result, por: str = "causa") -> pd.DataFrame:
    """Contagem de variáveis excluídas por causa (a leitura executiva do funil).

    ``por``: ``"causa"`` (default — a causa curta da etapa, em
    :data:`CAUSAS_CURTAS`), ``"etapa"`` (só o rótulo da etapa) ou ``"motivo"``
    (o texto por extenso, com os números do caso). Devolve um DataFrame com
    ``causa``, ``etapa``, ``n`` e ``pct`` (% das candidatas), do maior para o
    menor."""
    tab = _tabela(result)
    n_cand = max(len(tab), 1)
    exc = tab[tab["decisao"].astype(str) == "excluida"] if len(tab) else tab
    linhas = {}
    for r in exc.itertuples(index=False):
        etapa = str(getattr(r, "etapa_saida", "") or "")
        if por == "etapa":
            chave = rotulo_etapa(etapa) if etapa else "sem etapa registrada"
        elif por == "motivo":
            chave = _curto(getattr(r, "motivo", "") or "sem motivo registrado", 70)
        else:
            chave = (CAUSAS_CURTAS.get(etapa)
                     or (rotulo_etapa(etapa) if etapa else "sem etapa registrada"))
        item = linhas.setdefault(chave, {"causa": chave, "etapa": etapa, "n": 0})
        item["n"] += 1
    out = pd.DataFrame(list(linhas.values()), columns=["causa", "etapa", "n"])
    if len(out):
        out["pct"] = 100.0 * out["n"] / n_cand
        out = out.sort_values("n", ascending=False, kind="stable").reset_index(drop=True)
    else:
        out["pct"] = pd.Series(dtype="float64")
    return out


# ======================================================================
# Gráficos
# ======================================================================
def plot_funil(result, figsize=None, dpi=150, save_path=None, ax=None):
    """**Funil da seleção**: quantas variáveis restam depois de cada etapa.

    Barras horizontais decrescentes, de ``candidatas`` até ``selecionadas``,
    rotuladas com o número de sobreviventes e com quantas caíram na etapa
    (``−k``, em vermelho). É o gráfico-síntese da apresentação: em uma imagem
    conta de onde partimos, o que cada régua custou e onde paramos."""
    fun = _funil(result)
    cont = _por_decisao(result)
    n_sel, n_rev = cont["selecionada"], cont["revisar"]

    linhas = []
    for r in fun.itertuples(index=False):
        etapa = str(getattr(r, "etapa", ""))
        linhas.append({"rotulo": rotulo_etapa(etapa),
                       "n": int(getattr(r, "n_saida", 0) or 0),
                       "perda": int(getattr(r, "n_excluidas", 0) or 0),
                       "inicio": etapa == "candidatas"})
    if not linhas:                                  # resultado sem funil
        n_cand = len(_tabela(result))
        linhas.append({"rotulo": "candidatas", "n": n_cand, "perda": 0,
                       "inicio": True})
    linhas.append({"rotulo": "selecionadas", "n": n_sel, "perda": 0, "inicio": False,
                   "fim": True})

    if figsize is None:
        figsize = (9.0, max(2.8, 0.52 * len(linhas) + 1.3))
    fig, ax = _new_ax(figsize, dpi, ax)
    n_cand = linhas[0]["n"]
    if n_cand <= 0:
        return _aviso_no_eixo(fig, ax, "nenhuma variável candidata na seleção",
                              save_path, dpi)

    cores = _degrade(max(len(linhas) - 1, 1))
    y = np.arange(len(linhas))[::-1]                # primeira linha no topo
    for i, item in enumerate(linhas):
        cor = (DECISAO_COR["selecionada"] if item.get("fim")
               else cores[min(i, len(cores) - 1)])
        ax.barh(y[i], item["n"], color=cor, edgecolor="white", height=0.66)
        pct = 100.0 * item["n"] / n_cand
        ax.annotate(f"{item['n']}  ({pct:.0f}%)", (item["n"], y[i]),
                    textcoords="offset points", xytext=(6, 0), ha="left",
                    va="center", fontsize=9, fontweight="bold", color=_AZUL)
        if item["perda"]:
            ax.annotate(f"−{item['perda']}", (item["n"], y[i]),
                        textcoords="offset points", xytext=(74, 0), ha="left",
                        va="center", fontsize=8.5, fontweight="bold",
                        color=DECISAO_COR["excluida"])
    ax.set_yticks(y)
    ax.set_yticklabels([_curto(i["rotulo"], 26) for i in linhas], fontsize=9.5)
    ax.set_xlim(0, n_cand * (1.42 if any(i["perda"] for i in linhas) else 1.24))
    ax.set_xlabel("variáveis que seguem")
    _eixo_inteiro(ax)
    ax.grid(axis="x", alpha=0.12)
    for sp in ("top", "right", "left"):
        ax.spines[sp].set_visible(False)
    extra = f" · {n_rev} a revisar" if n_rev else ""
    ax.set_title(f"Funil da seleção de variáveis · {_problema(result)}\n"
                 f"{n_cand} candidatas → {n_sel} selecionadas{extra}",
                 fontsize=11, fontweight="bold", color=_AZUL)
    return _finish(fig, save_path, dpi)


def plot_motivos(result, top=10, por="causa", figsize=None, dpi=150, save_path=None,
                 ax=None):
    """**Por que perdemos variáveis**: contagem de exclusões por causa.

    A leitura executiva do funil — o funil diz *quantas* saíram em cada etapa,
    este diz *por quê*. ``por`` escolhe a granularidade (``"causa"``,
    ``"etapa"`` ou ``"motivo"`` por extenso) e ``top`` limita o número de barras
    (o resto entra numa barra ``outras causas``)."""
    d = causas_exclusao(result, por=por)
    n_cand = len(_tabela(result))
    if figsize is None:
        figsize = (9.0, max(2.8, 0.46 * max(len(d), 1) + 1.6))
    fig, ax = _new_ax(figsize, dpi, ax)
    if not len(d):
        return _aviso_no_eixo(
            fig, ax, "nenhuma variável foi excluída — todas as candidatas "
                     "sobreviveram às etapas executadas", save_path, dpi)
    if top and len(d) > int(top):
        resto = d.iloc[int(top):]
        outras = pd.DataFrame([{"causa": f"outras causas ({len(resto)})", "etapa": "",
                                "n": int(resto["n"].sum()),
                                "pct": float(resto["pct"].sum())}])
        d = pd.concat([d.iloc[:int(top)], outras], ignore_index=True)
    n_exc = int(d["n"].sum())
    y = np.arange(len(d))[::-1]
    cores = _degrade(len(d), inicio=DECISAO_COR["excluida"], fim="#e0a79e")
    ax.barh(y, d["n"].to_numpy(dtype="float64"), color=cores, edgecolor="white",
            height=0.68)
    for yi, n, pct in zip(y, d["n"], d["pct"]):
        ax.annotate(f"{int(n)}  ({pct:.0f}% das candidatas)", (n, yi),
                    textcoords="offset points", xytext=(6, 0), ha="left",
                    va="center", fontsize=8.5, color=_AZUL)
    ax.set_yticks(y)
    ax.set_yticklabels([_curto(c, 46) for c in d["causa"]], fontsize=9)
    ax.set_xlim(0, max(float(d["n"].max()), 1.0) * 1.75)
    ax.set_xlabel("variáveis excluídas")
    _eixo_inteiro(ax)
    ax.grid(axis="x", alpha=0.12)
    for sp in ("top", "right", "left"):
        ax.spines[sp].set_visible(False)
    ax.set_title(f"Por que as variáveis saíram · {n_exc} de {n_cand} candidatas",
                 fontsize=11, fontweight="bold", color=_AZUL)
    return _finish(fig, save_path, dpi)


def plot_iv_ranking(result, top=20, figsize=None, dpi=150, save_path=None, ax=None):
    """**Ranking de IV**, colorido pela decisão, com o corte usado na esteira.

    Barras horizontais do IV de cada variável (as que chegaram à etapa de IV),
    da maior para a menor, pintadas por ``selecionada``/``a revisar``/
    ``excluída`` e cortadas pela linha tracejada do ``min_iv`` efetivo (lido da
    política). ``top`` limita o número de barras — quantas ficaram de fora é
    anotado no canto."""
    tab = _tabela(result)
    if len(tab):
        d = tab[["variavel", "rotulo", "decisao", "iv"]].copy()
        d["iv"] = pd.to_numeric(d["iv"], errors="coerce")
        d = d[d["iv"].notna()].sort_values("iv", ascending=False, kind="stable")
    else:
        d = pd.DataFrame(columns=["variavel", "rotulo", "decisao", "iv"])
    n_total = len(d)
    fora = 0
    if top and n_total > int(top):
        fora = n_total - int(top)
        d = d.head(int(top))
    if figsize is None:
        figsize = (8.6, max(2.8, 0.36 * max(len(d), 1) + 1.6))
    fig, ax = _new_ax(figsize, dpi, ax)
    if not len(d):
        return _aviso_no_eixo(
            fig, ax, "nenhuma variável chegou à etapa de IV — sem poder "
                     "discriminante medido nesta execução", save_path, dpi)

    min_iv = _f(_params(result).get("min_iv"))
    y = np.arange(len(d))[::-1]
    cores = [DECISAO_COR.get(str(v), _CINZA) for v in d["decisao"]]
    ax.barh(y, d["iv"].to_numpy(dtype="float64"), color=cores, edgecolor="white",
            height=0.7)
    for yi, iv in zip(y, d["iv"]):
        ax.annotate(_fmt_num(iv, 3), (float(iv), yi), textcoords="offset points",
                    xytext=(5, 0), ha="left", va="center", fontsize=8, color=_AZUL)
    ax.set_yticks(y)
    ax.set_yticklabels([_curto(r, 32) for r in d["rotulo"]], fontsize=8.5)
    ax.set_xlabel("IV (information value)")
    ax.grid(axis="x", alpha=0.12)
    for sp in ("top", "right", "left"):
        ax.spines[sp].set_visible(False)
    iv_max = float(d["iv"].max())
    ax.set_xlim(0, max(iv_max, min_iv or 0.0) * 1.22 + 1e-6)
    if min_iv is not None:
        ax.axvline(min_iv, color=_AZUL, ls="--", lw=1.2, alpha=0.75,
                   label=f"mínimo {_fmt_num(min_iv, 3)}")

    from matplotlib.patches import Patch
    presentes = [dec for dec in ORDEM_DECISAO if (d["decisao"] == dec).any()]
    handles = [Patch(facecolor=DECISAO_COR[dec], label=DECISAO_LABEL[dec])
               for dec in presentes]
    if min_iv is not None:
        from matplotlib.lines import Line2D
        handles.append(Line2D([0], [0], color=_AZUL, ls="--", lw=1.2,
                              label=f"corte de IV = {_fmt_num(min_iv, 3)}"))
    if handles:
        ax.legend(handles=handles, fontsize=8, loc="lower right", framealpha=0.9)
    if fora:
        ax.annotate(f"+{fora} variável(is) fora do top {int(top)}", (0.99, 1.01),
                    xycoords="axes fraction", ha="right", va="bottom", fontsize=8,
                    color=_CINZA)
    ax.set_title(f"Poder discriminante por variável · {_problema(result)}",
                 fontsize=11, fontweight="bold", color=_AZUL)
    return _finish(fig, save_path, dpi)


def plot_iv_psi(result, annotate_top=8, figsize=(7.8, 5.6), dpi=150, save_path=None,
                ax=None):
    """**IV × PSI em quadrantes**: poder discriminante contra estabilidade.

    Dispersão com as linhas de corte da política (``min_iv`` no eixo x e
    ``max_psi`` no eixo y) dividindo o plano em quatro leituras — *forte e
    estável* (o que queremos), *forte mas instável*, *fraca e estável* e
    *descartar*. Os pontos são coloridos pela decisão e as ``annotate_top``
    variáveis de maior IV recebem o nome (anotar todas deixaria o gráfico
    ilegível)."""
    tab = _tabela(result)
    fig, ax = _new_ax(figsize, dpi, ax)
    if len(tab):
        d = tab[["variavel", "rotulo", "decisao", "iv", "pior_psi"]].copy()
        d["iv"] = pd.to_numeric(d["iv"], errors="coerce")
        d["pior_psi"] = pd.to_numeric(d["pior_psi"], errors="coerce")
        d = d[d["iv"].notna() & d["pior_psi"].notna()]
    else:
        d = pd.DataFrame(columns=["variavel", "rotulo", "decisao", "iv", "pior_psi"])
    if not len(d):
        return _aviso_no_eixo(
            fig, ax, "sem PSI calculado — é preciso ao menos uma amostra de "
                     "comparação além da referência e a etapa 'psi' na esteira",
            save_path, dpi)

    par = _params(result)
    min_iv = _f(par.get("min_iv")) or 0.0
    max_psi = _f(par.get("max_psi")) or 0.25

    # Uma variável com PSI explosivo (deslocada de safra) achataria todo o resto
    # contra o eixo: o eixo vai até 3× o corte e quem passa disso é desenhado no
    # topo como ▲ (a leitura do gráfico é a vizinhança dos cortes, não a escala).
    psi = d["pior_psi"].to_numpy(dtype="float64")
    teto = max_psi * 3.0
    y_dado = max(float(psi.max()) * 1.12 if float(psi.max()) <= teto else teto,
                 max_psi * 1.5, 1e-3)
    d = d.assign(psi_plot=np.minimum(psi, y_dado), acima=psi > y_dado)
    n_fora = int(d["acima"].sum())
    for dec in ORDEM_DECISAO:
        sub = d[d["decisao"] == dec]
        if not len(sub):
            continue
        dentro, fora = sub[~sub["acima"]], sub[sub["acima"]]
        if len(dentro):
            ax.scatter(dentro["iv"], dentro["psi_plot"], s=70, c=DECISAO_COR[dec],
                       alpha=0.82, edgecolor="white", linewidth=0.9, zorder=3,
                       label=DECISAO_LABEL[dec])
        if len(fora):
            ax.scatter(fora["iv"], fora["psi_plot"], s=105, marker="^",
                       c=DECISAO_COR[dec], alpha=0.82, edgecolor="white",
                       linewidth=0.9, zorder=3,
                       label=DECISAO_LABEL[dec] if not len(dentro) else None)
    x_max = max(float(d["iv"].max()) * 1.18, min_iv * 2.2, 1e-3)
    ax.set_xlim(0, x_max)
    ax.set_ylim(0, y_dado * 1.28)          # faixa livre no topo para a legenda
    ax.axvline(min_iv, color="#bbb", ls="--", lw=1.1, zorder=1)
    ax.axhline(max_psi, color="#bbb", ls="--", lw=1.1, zorder=1)

    # rótulos dos quadrantes (canto de cada um, discretos)
    quadrantes = [
        (0.985, 0.045, "right", "bottom", "forte e estável"),
        (0.985, 0.855, "right", "top", "forte mas instável"),
        (0.015, 0.045, "left", "bottom", "fraca e estável"),
        (0.015, 0.855, "left", "top", "descartar"),
    ]
    for fx, fy, ha, va, txt in quadrantes:
        ax.text(fx, fy, txt, transform=ax.transAxes, ha=ha, va=va, fontsize=8.5,
                color=_CINZA, style="italic", zorder=2,
                bbox=dict(facecolor="white", alpha=0.7, edgecolor="none", pad=1.5))

    if annotate_top:
        topo = d.sort_values("iv", ascending=False).head(int(annotate_top))
        y_lo, y_hi = ax.get_ylim()
        span = max(y_hi - y_lo, 1e-12)
        for k, (_, r) in enumerate(topo.iterrows()):
            xv, yv = float(r["iv"]), float(r["psi_plot"])
            esquerda = xv > 0.72 * x_max                   # não estourar a direita
            alto = (yv - y_lo) / span
            # perto das bordas o rótulo vai sempre para dentro; a alternância em
            # duas alturas separa nomes de pontos praticamente colados
            if alto < 0.12:
                dy = 9 if k % 2 == 0 else 21
            elif alto > 0.80:
                dy = -13 if k % 2 == 0 else -25
            else:
                dy = 9 if k % 2 else -13
            ax.annotate(_curto(r["rotulo"], 22), (xv, yv), textcoords="offset points",
                        xytext=(-8 if esquerda else 8, dy),
                        ha="right" if esquerda else "left", fontsize=7.5,
                        color=_AZUL, zorder=4)
    ax.set_xlabel(f"IV (corte {_fmt_num(min_iv, 3)})")
    extra = (f"; ▲ = acima de {_fmt_num(y_dado, 2)}" if n_fora else "")
    ax.set_ylabel(f"pior PSI entre amostras (máximo {_fmt_num(max_psi, 2)}{extra})")
    ax.grid(alpha=0.12)
    ax.legend(fontsize=8, loc="upper center", ncol=3, framealpha=0.9)
    ax.set_title(f"Poder × estabilidade · {_problema(result)}", fontsize=11,
                 fontweight="bold", color=_AZUL)
    return _finish(fig, save_path, dpi)


# ======================================================================
# Tabelas de apresentação
# ======================================================================
_ORDEM_RANK = {"selecionada": 0, "revisar": 1, "excluida": 2}


def _ordena_tabela(d: pd.DataFrame) -> pd.DataFrame:
    """Ordena por decisão (selecionada → revisar → excluída) e IV decrescente."""
    if not len(d):
        return d
    rank = d["decisao"].astype(str).map(_ORDEM_RANK)
    iv = pd.to_numeric(d["iv"], errors="coerce")
    aux = d.assign(_rank=rank.fillna(9).to_numpy(),
                   _iv=iv.fillna(-np.inf).to_numpy())
    aux = aux.sort_values(["_rank", "_iv"], ascending=[True, False], kind="stable")
    return aux.drop(columns=["_rank", "_iv"])


def tabela_decisoes(result, numerico: bool = False, decisao=None, ordenar: bool = True):
    """Tabela de decisões **pronta para a apresentação** (uma linha por candidata).

    Colunas renomeadas para rótulos legíveis (``Variável``, ``Tipo``, ``Decisão``,
    ``Etapa``, ``Motivo``, ``IV``, ``PSI (pior)``…), IV/PSI com 3 casas e vírgula
    decimal, faltantes como ``"12,3%"`` e ausências como ``"—"``.

    Parameters
    ----------
    numerico:
        ``True`` devolve os **valores crus** (as colunas canônicas de
        ``SelectionResult.tabela``, com os dtypes originais) — é a versão para
        Excel/análise.
    decisao:
        Filtra por uma decisão (``"selecionada"``/``"revisar"``/``"excluida"``)
        ou por uma lista delas. ``None`` traz todas.
    ordenar:
        ``True`` (default) ordena por decisão e IV decrescente — a leitura
        natural do relatório. ``False`` mantém a ordem de entrada das candidatas.
    """
    d = _tabela(result)
    if decisao is not None:
        alvos = [decisao] if isinstance(decisao, str) else list(decisao)
        d = d[d["decisao"].astype(str).isin([str(a) for a in alvos])]
    if ordenar:
        d = _ordena_tabela(d)
    d = d.reset_index(drop=True)
    if numerico:
        return d
    if not len(d):
        return pd.DataFrame(columns=[ROTULOS_TABELA[c] for c in COLUNAS_TABELA])
    out = pd.DataFrame({
        "Variável": d["variavel"].astype(str),
        "Rótulo": d["rotulo"].astype(str),
        "Tipo": [_TIPO_LABEL.get(str(v), str(v)) for v in d["tipo"]],
        "Decisão": [DECISAO_LABEL.get(str(v), str(v)) for v in d["decisao"]],
        "Etapa": [rotulo_etapa(v) if str(v) else "—" for v in d["etapa_saida"]],
        "Motivo": d["motivo"].astype(str),
        "IV": [_fmt_num(v, 4) for v in d["iv"]],
        "Força": [str(v) for v in d["forca"]],
        "PSI (pior)": [_fmt_num(v, 3) for v in d["pior_psi"]],
        "Estabilidade": [str(v) for v in d["estabilidade"]],
        "Tendência": [str(v) for v in d["tendencia"]],
        "Inversões": [_fmt_int(v) for v in d["n_inversoes"]],
        "Faltantes": [_fmt_pct(v, 1) for v in d["missing_pct"]],
        "Categorias": [_fmt_int(v) for v in d["n_categorias"]],
    })
    return out


def tabela_funil(result, numerico: bool = False):
    """Tabela do funil **pronta para a apresentação** (uma linha por etapa).

    Acrescenta ``% das candidatas`` (quanto do conjunto inicial ainda seguia
    depois da etapa) e traduz a chave da etapa para o rótulo em pt-BR.
    ``numerico=True`` devolve os valores crus (colunas canônicas + ``retido_pct``
    em 0–100)."""
    d = _funil(result)
    n_cand = int(d["n_saida"].iloc[0]) if len(d) else len(_tabela(result))
    base = max(n_cand, 1)
    d = d.reset_index(drop=True)
    retido = [100.0 * int(v) / base for v in d["n_saida"]] if len(d) else []
    if numerico:
        out = d.copy()
        out["retido_pct"] = pd.Series(retido, dtype="float64")
        return out
    if not len(d):
        return pd.DataFrame(columns=[ROTULOS_FUNIL[c] for c in COLUNAS_FUNIL]
                                    + [ROTULOS_FUNIL["retido_pct"]])
    return pd.DataFrame({
        "Etapa": [rotulo_etapa(v) for v in d["etapa"]],
        "Entraram": [_fmt_int(v) for v in d["n_entrada"]],
        "Excluídas": [_fmt_int(v) for v in d["n_excluidas"]],
        "A revisar": [_fmt_int(v) for v in d["n_revisar"]],
        "Seguiram": [_fmt_int(v) for v in d["n_saida"]],
        "% das candidatas": [_fmt_pct(v, 0) for v in retido],
    })


def _valor_politica(v) -> str:
    if v is None:
        return "—"
    if isinstance(v, bool):
        return "sim" if v else "não"
    if isinstance(v, (list, tuple)):
        return ", ".join(str(x) for x in v) if len(v) else "—"
    if isinstance(v, float):
        return _fmt_num(v, 4)
    return str(v)


def tabela_politica(result):
    """**Política efetiva** da execução em duas colunas (``Item`` / ``Valor``).

    É o bloco de auditoria/reprodutibilidade do relatório: alvo, amostra de
    referência, etapas na ordem executada, se foi aplicado no segmentador, os
    avisos e **todos** os parâmetros já resolvidos (nada implícito)."""
    pol = _politica(result)
    linhas = [
        ("Alvo (coluna)", _valor_politica(pol.get("alvo"))),
        ("Problema (rótulo do alvo)", _valor_politica(pol.get("problema"))),
        ("Tipo de tarefa", _valor_politica(pol.get("task_type"))),
        ("Amostra de referência", _valor_politica(pol.get("amostra_referencia"))),
        ("Etapas executadas", " → ".join(rotulo_etapa(e)
                                         for e in (pol.get("etapas") or [])) or "—"),
        ("Aplicado no segmentador", _valor_politica(pol.get("aplicado"))),
        ("Versão da política", _valor_politica(pol.get("versao"))),
    ]
    for aviso in (pol.get("avisos") or []):
        linhas.append(("Aviso", str(aviso)))
    for chave in sorted(_params(result)):
        linhas.append((f"parâmetro · {chave}", _valor_politica(_params(result)[chave])))
    return pd.DataFrame(linhas, columns=["Item", "Valor"])


# ======================================================================
# Relatório HTML autocontido
# ======================================================================
_CSS = """
:root{
  --bg:#ffffff; --fg:#1f2d3a; --muted:#6b7480; --accent:#0f3d57;
  --line:#e5e9ee; --surface:#f4f7f9; --shadow:0 1px 2px rgba(15,61,87,.06);
  --ok-tx:#0f6b45; --ok-bg:#eaf5ee; --warn-tx:#8a6100; --warn-bg:#fdf3e0;
  --bad-tx:#9c2f21; --bad-bg:#fbeceb;
}
@media (prefers-color-scheme: dark){
  :root{
    --bg:#11171c; --fg:#e3eaf0; --muted:#93a1ad; --accent:#7fb6d8;
    --line:#26313a; --surface:#18212a; --shadow:none;
    --ok-tx:#7fd1a8; --ok-bg:#122a20; --warn-tx:#e5be6a; --warn-bg:#2b2313;
    --bad-tx:#f0a094; --bad-bg:#2e1a17;
  }
}
*{box-sizing:border-box}
body{font-family:-apple-system,Segoe UI,Roboto,Arial,sans-serif;background:var(--bg);
  color:var(--fg);margin:0;padding:26px 22px 40px;line-height:1.45}
.wrap{max-width:1120px;margin:0 auto}
h1{font-size:22px;margin:0 0 4px;color:var(--accent)}
h2{font-size:15.5px;margin:30px 0 8px;color:var(--accent);
  border-bottom:2px solid var(--line);padding-bottom:5px}
p.sub{color:var(--muted);font-size:13px;margin:0 0 4px}
p.meta{color:var(--muted);font-size:11.5px;margin:0 0 18px}
p.nota{color:var(--muted);font-size:12px;margin:4px 0 10px}
.cards{display:flex;flex-wrap:wrap;gap:12px;margin:14px 0 6px}
.card{flex:1 1 150px;background:var(--surface);border:1px solid var(--line);
  border-radius:10px;padding:12px 14px;box-shadow:var(--shadow)}
.card .n{font-size:26px;font-weight:700;line-height:1.1}
.card .lab{font-size:11.5px;color:var(--muted);text-transform:uppercase;
  letter-spacing:.04em;margin-top:2px}
.card.ok .n{color:var(--ok-tx)} .card.warn .n{color:var(--warn-tx)}
.card.bad .n{color:var(--bad-tx)} .card.base .n{color:var(--accent)}
.causas{background:var(--surface);border:1px solid var(--line);border-radius:10px;
  padding:12px 16px;margin:12px 0 4px}
.causas ol{margin:6px 0 0;padding-left:20px;font-size:13px}
.causas li{margin:3px 0}
.banner{background:var(--warn-bg);color:var(--warn-tx);border:1px solid var(--line);
  border-radius:8px;padding:9px 13px;font-size:12.5px;margin:12px 0}
.tbl-wrap{overflow-x:auto;margin:6px 0 4px}
table{border-collapse:collapse;font-size:12.5px;width:100%}
th,td{border:1px solid var(--line);padding:5px 9px;text-align:center;
  vertical-align:top}
th{background:var(--surface);font-weight:600;color:var(--accent);
  position:sticky;top:0}
td.l,th.l{text-align:left}
td.var,th.var{min-width:120px}
td.motivo,th.motivo{min-width:300px}
tbody tr:nth-child(even) td{background:color-mix(in srgb,var(--surface) 55%,transparent)}
.chip{display:inline-block;padding:1px 8px;border-radius:999px;font-size:11.5px;
  font-weight:600;white-space:nowrap}
.chip.selecionada{background:var(--ok-bg);color:var(--ok-tx)}
.chip.revisar{background:var(--warn-bg);color:var(--warn-tx)}
.chip.excluida{background:var(--bad-bg);color:var(--bad-tx)}
figure{margin:8px 0 4px}
figure img{max-width:100%;height:auto;border:1px solid var(--line);border-radius:8px;
  background:#fff}
figcaption{color:var(--muted);font-size:11.5px;margin-top:4px}
footer{color:var(--muted);font-size:11px;margin-top:34px;border-top:1px solid var(--line);
  padding-top:10px}
"""


def _esc(x) -> str:
    return _html.escape("" if x is None else str(x))


def _tbl_html(df, left=(), classe="") -> str:
    """DataFrame de textos → tabela HTML (colunas em ``left`` alinhadas à esquerda)."""
    if df is None or not len(df):
        return "<p class='nota'>— sem dados —</p>"
    cols = list(df.columns)
    head = "".join(f"<th class='{'l' if c in left else ''}'>{_esc(c)}</th>"
                   for c in cols)
    linhas = []
    for _, r in df.iterrows():
        tds = "".join(f"<td class='{'l' if c in left else ''}'>{_esc(r[c])}</td>"
                      for c in cols)
        linhas.append(f"<tr>{tds}</tr>")
    return (f"<div class='tbl-wrap'><table class='{classe}'><thead><tr>{head}</tr>"
            f"</thead><tbody>{''.join(linhas)}</tbody></table></div>")


def _tbl_decisoes_html(result) -> str:
    """Tabela completa de decisões, com a decisão em *chip* colorido."""
    fmt = tabela_decisoes(result)
    cru = tabela_decisoes(result, numerico=True)
    if not len(fmt):
        return "<p class='nota'>— sem variáveis candidatas —</p>"
    cols = ["Variável", "Tipo", "Decisão", "Etapa", "Motivo", "IV", "Força",
            "PSI (pior)", "Estabilidade", "Tendência", "Faltantes", "Categorias"]
    # o motivo é a coluna que carrega a narrativa: ganha largura mínima e a
    # tabela rola na horizontal em vez de espremê-la
    classe = {"Variável": "l var", "Etapa": "l", "Motivo": "l motivo"}
    head = "".join(f"<th class='{classe.get(c, '')}'>{_esc(c)}</th>" for c in cols)
    linhas = []
    for i in range(len(fmt)):
        dec = str(cru["decisao"].iloc[i])
        tds = []
        for c in cols:
            if c == "Decisão":
                tds.append(f"<td><span class='chip {_esc(dec)}'>"
                           f"{_esc(fmt[c].iloc[i])}</span></td>")
            else:
                tds.append(f"<td class='{classe.get(c, '')}'>"
                           f"{_esc(fmt[c].iloc[i])}</td>")
        linhas.append(f"<tr>{''.join(tds)}</tr>")
    return (f"<div class='tbl-wrap'><table><thead><tr>{head}</tr></thead>"
            f"<tbody>{''.join(linhas)}</tbody></table></div>")


def _figura_html(construtor, legenda: str, dpi: int) -> str:
    """Figura → ``<figure>`` com data URI. Falha de desenho vira nota, não erro."""
    try:
        fig = construtor()
        uri = fig_to_data_uri(fig, dpi=dpi)
    except Exception as exc:  # noqa: BLE001 - o relatório nunca cai por um gráfico
        return (f"<p class='nota'>gráfico não gerado "
                f"({_esc(type(exc).__name__)}: {_esc(exc)})</p>")
    return (f"<figure><img alt='{_esc(legenda)}' src='{uri}'/>"
            f"<figcaption>{_esc(legenda)}</figcaption></figure>")


def _contexto_seg(seg) -> list:
    """Frases de contexto extraídas do segmentador (quando ele é passado)."""
    if seg is None:
        return []
    partes = []
    alvo = getattr(seg, "problem_label", None) or getattr(seg, "target", None)
    if alvo:
        partes.append(f"alvo: {alvo}")
    try:
        partes.append(f"{len(seg.df):,}".replace(",", ".") + " linhas na base")
    except Exception:  # noqa: BLE001 - contexto é cosmético
        pass
    try:
        amostras = list(seg._samples())
        if amostras:
            partes.append("amostras: " + ", ".join(str(a) for a in amostras))
    except Exception:  # noqa: BLE001
        pass
    try:
        partes.append(f"{len(seg.candidates)} variáveis candidatas")
    except Exception:  # noqa: BLE001
        pass
    return partes


def build_selection_report_html(result, seg=None, title=None, subtitle=None,
                                top_iv: int = 20, annotate_top: int = 8,
                                dpi: int = 110, incluir_graficos: bool = True) -> str:
    """Monta o **relatório de seleção** como uma página HTML autocontida.

    Uma página única — CSS embutido, figuras em ``data:image/png;base64`` e
    nenhuma requisição externa —, portanto abrível fora do notebook e anexável à
    apresentação. Seções: cabeçalho, **sumário executivo** em cartões (candidatas,
    selecionadas, excluídas, a revisar e as 3 principais causas de exclusão),
    **funil** (gráfico + tabela), **por que perdemos variáveis**, **ranking de
    IV**, **IV × PSI**, a **tabela completa de decisões** com o motivo por
    extenso e a **política** usada.

    Parameters
    ----------
    result:
        :class:`~yggdrasil.credit_risk.model.selection.SelectionResult`.
    seg:
        Segmentador (opcional) — quando passado, enriquece o cabeçalho com o
        rótulo do alvo, as amostras e o tamanho da base. Sem ele, o relatório sai
        completo apenas com o que está no ``result``.
    title, subtitle:
        Sobrescrevem o título e o subtítulo padrão.
    top_iv, annotate_top, dpi:
        Repassados aos gráficos (nº de barras no ranking de IV, nº de nomes
        anotados no IV × PSI e resolução das imagens embutidas).
    incluir_graficos:
        ``False`` gera só o texto e as tabelas (relatório mais leve).
    """
    pol = _politica(result)
    cont = _por_decisao(result)
    n_cand = len(_tabela(result))
    n_sel, n_rev, n_exc = cont["selecionada"], cont["revisar"], cont["excluida"]
    problema = _problema(result)

    ttl = title or f"Relatório de seleção de variáveis · {problema}"
    partes_sub = [f"{n_cand} candidatas → {n_sel} selecionadas"]
    if n_rev:
        partes_sub.append(f"{n_rev} a revisar")
    partes_sub.append(f"{n_exc} excluídas")
    if pol.get("amostra_referencia"):
        partes_sub.append(f"amostra de referência: {pol['amostra_referencia']}")
    sub = subtitle or " · ".join(partes_sub)
    contexto = _contexto_seg(seg)
    etapas = " → ".join(rotulo_etapa(e) for e in (pol.get("etapas") or [])) or "nenhuma"
    quando = datetime.now().strftime("%d/%m/%Y %H:%M")
    meta = [f"gerado em {quando}", f"etapas: {etapas}"]
    if contexto:
        meta.extend(contexto)

    top_causas = causas_exclusao(result).head(3)
    itens = "".join(f"<li><b>{int(r.n)}</b> variável(is) — {_esc(r.causa)} "
                    f"({r.pct:.0f}% das candidatas)</li>"
                    for r in top_causas.itertuples(index=False))
    causas_html = (f"<div class='causas'><b>Principais causas de exclusão</b>"
                   f"<ol>{itens}</ol></div>" if itens else
                   "<p class='nota'>Nenhuma variável foi excluída nesta execução.</p>")

    banner = ""
    if not pol.get("aplicado", True):
        banner = ("<div class='banner'>Simulação: as decisões abaixo <b>não</b> "
                  "foram aplicadas no segmentador (nada foi incluído, excluído ou "
                  "categorizado).</div>")
    for aviso in (pol.get("avisos") or []):
        banner += f"<div class='banner'>Aviso: {_esc(aviso)}</div>"

    cards = (
        f"<div class='cards'>"
        f"<div class='card base'><div class='n'>{n_cand}</div>"
        f"<div class='lab'>candidatas</div></div>"
        f"<div class='card ok'><div class='n'>{n_sel}</div>"
        f"<div class='lab'>selecionadas</div></div>"
        f"<div class='card warn'><div class='n'>{n_rev}</div>"
        f"<div class='lab'>a revisar</div></div>"
        f"<div class='card bad'><div class='n'>{n_exc}</div>"
        f"<div class='lab'>excluídas</div></div>"
        f"</div>")

    def _fig(nome, construtor, legenda):
        if not incluir_graficos:
            return ""
        return _figura_html(construtor, legenda, dpi)

    corpo = [
        f"<h1>{_esc(ttl)}</h1>",
        f"<p class='sub'>{_esc(sub)}</p>",
        f"<p class='meta'>{_esc(' · '.join(meta))}</p>",
        banner,
        "<h2 id='sumario'>Sumário executivo</h2>",
        cards,
        causas_html,
        "<h2 id='funil'>Funil da seleção</h2>",
        "<p class='nota'>Quantas variáveis seguiram depois de cada etapa — o "
        "custo de cada régua sobre o conjunto inicial.</p>",
        _fig("funil", lambda: plot_funil(result), "Funil da seleção de variáveis"),
        _tbl_html(tabela_funil(result), left={"Etapa"}),
        "<h2 id='motivos'>Por que perdemos variáveis</h2>",
        "<p class='nota'>A leitura executiva do funil: quantas variáveis cada "
        "causa levou embora.</p>",
        _fig("motivos", lambda: plot_motivos(result),
             "Variáveis excluídas por causa"),
        "<h2 id='iv'>Ranking de poder discriminante (IV)</h2>",
        "<p class='nota'>IV por variável, colorido pela decisão; a linha "
        "tracejada é o corte de IV efetivamente usado.</p>",
        _fig("iv", lambda: plot_iv_ranking(result, top=top_iv),
             "IV por variável, colorido pela decisão"),
        "<h2 id='iv-psi'>Poder × estabilidade (IV × PSI)</h2>",
        "<p class='nota'>Quadrantes formados pelos cortes de IV e de PSI da "
        "política: o canto inferior direito é o território desejável (sinal "
        "forte e distribuição estável).</p>",
        _fig("iv_psi", lambda: plot_iv_psi(result, annotate_top=annotate_top),
             "IV × pior PSI, com os quadrantes de decisão"),
        "<h2 id='decisoes'>Decisões por variável</h2>",
        "<p class='nota'>Uma linha por candidata, com a etapa de saída e o "
        "motivo por extenso.</p>",
        _tbl_decisoes_html(result),
        "<h2 id='politica'>Política usada</h2>",
        "<p class='nota'>Parâmetros efetivos da execução — o bloco de "
        "reprodutibilidade e auditoria.</p>",
        _tbl_html(tabela_politica(result), left={"Item", "Valor"}),
        "<footer>Gerado por yggdrasil.credit_risk.model — esteira de seleção de "
        "variáveis. Decisões automáticas são um ponto de partida: a validação "
        "final da lista é do time de modelagem.</footer>",
    ]
    return ("<!doctype html><html lang='pt-BR'><head><meta charset='utf-8'>"
            "<meta name='viewport' content='width=device-width,initial-scale=1'>"
            f"<title>{_esc(ttl)}</title><style>{_CSS}</style></head><body>"
            f"<div class='wrap'>{''.join(corpo)}</div></body></html>")


def save_selection_report(result, path, seg=None, **kwargs) -> str:
    """Grava :func:`build_selection_report_html` em ``path`` e devolve o caminho.

    O arquivo é autocontido: basta abri-lo no navegador (ou anexá-lo ao
    material da apresentação)."""
    html_doc = build_selection_report_html(result, seg=seg, **kwargs)
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(html_doc)
    return str(path)


# ======================================================================
# Excel (openpyxl é OPCIONAL — import lazy)
# ======================================================================
def export_selection_xlsx(result, path) -> str:
    """Exporta a seleção para um Excel multi-abas (``.xlsx``) em ``path``.

    Abas: **Decisoes** (uma linha por candidata, com os valores **numéricos** e
    cabeçalhos legíveis), **Funil** (uma linha por etapa) e **Politica** (os
    parâmetros efetivos, para auditoria).

    Requer o pacote OPCIONAL **openpyxl** — o import é lazy (só aqui) e, sem ele
    instalado, sobe um :class:`ImportError` com instrução amigável (a biblioteca
    não ganha dependência nova por causa do relatório).

    Retorna o próprio ``path``."""
    try:
        import openpyxl  # noqa: F401  — dependência OPCIONAL (import lazy)
    except ImportError as e:
        raise ImportError(
            "A exportação para Excel requer o pacote opcional 'openpyxl' "
            "(instale com: pip install openpyxl).") from e
    from openpyxl.utils import get_column_letter

    dec = tabela_decisoes(result, numerico=True).rename(columns=ROTULOS_TABELA)
    dec["Tipo"] = [_TIPO_LABEL.get(str(v), str(v)) for v in dec["Tipo"]]
    dec["Decisão"] = [DECISAO_LABEL.get(str(v), str(v)) for v in dec["Decisão"]]
    dec["Etapa"] = [rotulo_etapa(v) if str(v) else "" for v in dec["Etapa"]]
    # faltantes vêm em 0–100; no Excel viram fração com formato de %
    dec["Faltantes"] = pd.to_numeric(dec["Faltantes"], errors="coerce") / 100.0
    fun = tabela_funil(result, numerico=True).rename(columns=ROTULOS_FUNIL)
    fun["Etapa"] = [rotulo_etapa(v) for v in fun["Etapa"]]
    fun["% das candidatas"] = pd.to_numeric(fun["% das candidatas"],
                                            errors="coerce") / 100.0
    pol = tabela_politica(result)

    pct_cols = {"Decisoes": ["Faltantes"], "Funil": ["% das candidatas"],
                "Politica": []}
    larguras = {"Motivo": 78, "Rótulo": 26, "Variável": 24, "Item": 32, "Valor": 60}

    with pd.ExcelWriter(path, engine="openpyxl") as writer:

        def _write(nome, df):
            df.to_excel(writer, sheet_name=nome, index=False)
            ws = writer.sheets[nome]
            ws.freeze_panes = "A2"                      # congela o cabeçalho
            idx = {str(c): i + 1 for i, c in enumerate(df.columns)}
            for col in pct_cols.get(nome, []):
                if col not in idx:
                    continue
                letra = get_column_letter(idx[col])
                for row in range(2, len(df) + 2):
                    ws[f"{letra}{row}"].number_format = "0.0%"
            for col, larg in larguras.items():
                if col in idx:
                    ws.column_dimensions[get_column_letter(idx[col])].width = larg

        _write("Decisoes", dec)
        _write("Funil", fun)
        _write("Politica", pol)
    return str(path)


__all__ = [
    "plot_funil", "plot_motivos", "plot_iv_ranking", "plot_iv_psi",
    "tabela_decisoes", "tabela_funil", "tabela_politica", "causas_exclusao",
    "build_selection_report_html", "save_selection_report", "export_selection_xlsx",
    "fig_to_data_uri", "DECISAO_LABEL", "DECISAO_COR", "ORDEM_DECISAO",
    "ROTULOS_TABELA", "ROTULOS_FUNIL", "CAUSAS_CURTAS",
]
