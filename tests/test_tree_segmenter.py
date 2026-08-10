"""
Testes da árvore de segmentação unificada ``yggdrasil.credit_risk.tree``.

Uma só classe (:class:`TreeSegmenter`) atende **classificação** (PD, alvo binário)
e **regressão** (LGD, alvo contínuo) via ``task_type``. A maioria dos testes é
**parametrizada nos dois tipos** (prova de que a unificação preserva o
comportamento das antigas classes separadas); os checks de valor específicos de
cada tarefa (KS/AUC vs MAE/RMSE; IV WoE vs IV contínuo) ficam isolados.
"""
from __future__ import annotations

import warnings

import numpy as np
import pandas as pd
import pytest

from yggdrasil.credit_risk.tree import TreeSegmenter

TASKS = ["classification", "regression"]


# ----------------------------------------------------------------------
# Geradores de dados: alvo binário (clf) ou contínuo (reg), mesmo desenho
# de features (score/ltv numérico + garantia categórica), com DES/OOT e safra.
# ----------------------------------------------------------------------
def make_df(task, n=4000, seed=0, com_na=False, com_oot=False):
    rng = np.random.default_rng(seed)
    x = rng.beta(2.5, 3, n) * 1.4 + 0.3                       # feature numérica
    gar = rng.choice(["A", "B", "C", "D"], n, p=[0.5, 0.22, 0.18, 0.1]).astype(object)
    if com_na:
        x[rng.random(n) < 0.08] = np.nan
        gar[rng.random(n) < 0.06] = np.nan
    lg = {"A": 0.0, "B": 0.10, "C": 0.16, "D": 0.30}
    risco = (0.1 + 0.4 * np.nan_to_num(x - 0.5, nan=0.35)
             + np.array([lg.get(g, 0.2) for g in gar]))
    if task == "classification":
        p = np.clip(risco, 0.01, 0.95)
        target = (rng.uniform(0, 1, n) < p).astype(float)
    else:
        target = np.clip(risco + rng.normal(0, 0.07, n), 0, 1)
    df = pd.DataFrame({"score": x, "garantia": gar, "target": target})
    if com_oot:
        meses = pd.date_range("2023-01-01", periods=10, freq="MS")
        df["dt_ref"] = rng.choice(meses, size=n)
        df["amostra"] = np.where(df["dt_ref"] >= meses[7], "OOT", "DES")
    else:
        df["amostra"] = "DES"
    return df


def _mk(task, **kw):
    df = kw.pop("df", None)
    if df is None:
        df = make_df(task, **{k: kw.pop(k) for k in ("n", "seed", "com_na", "com_oot")
                              if k in kw})
    return TreeSegmenter(df, target="target", task_type=task,
                         sample_col="amostra", ref_sample="DES", verbose=False, **kw)


@pytest.fixture(params=TASKS)
def task(request):
    return request.param


@pytest.fixture
def seg(task):
    return _mk(task)


def _cobertura_total(seg):
    cob = sum(s["mask"].sum() for s in seg.segments.values() if s["is_leaf"])
    return cob == len(seg.df)


# ----------------------------------------------------------------------
# Construção / validação
# ----------------------------------------------------------------------
def test_import_pacote():
    import yggdrasil
    from yggdrasil.credit_risk import TreeSegmenter as C
    assert C is TreeSegmenter
    assert isinstance(yggdrasil.__version__, str)


def test_task_type_invalido():
    df = make_df("classification", n=200)
    with pytest.raises(ValueError, match="task_type"):
        TreeSegmenter(df, target="target", task_type="binario", verbose=False)


def test_target_ausente_erro():
    df = make_df("classification", n=200).drop(columns=["target"])
    with pytest.raises(ValueError, match="alvo"):
        TreeSegmenter(df, target="target", task_type="classification", verbose=False)


def test_construtor_sample_col_ausente_erro(task):
    df = make_df(task, n=300)
    with pytest.raises(ValueError):
        TreeSegmenter(df, target="target", task_type=task,
                      sample_col="nao_existe", verbose=False)


# ----------------------------------------------------------------------
# Crescimento (manual num/cat, automático) + cobertura
# ----------------------------------------------------------------------
def test_grow_numerico(seg):
    seg.grow("score", splits=[0.8])
    assert len(seg.leaves()) >= 2
    assert _cobertura_total(seg)


def test_grow_categorico(seg):
    seg.grow("garantia", splits=[["A"], ["B", "C"], ["D"]])
    assert len(seg.leaves()) >= 2
    assert _cobertura_total(seg)


def test_grow_grupos_cat_repetidos_erro(seg):
    with pytest.raises(ValueError, match="repetida"):
        seg.grow("garantia", splits=[["A", "B"], ["B", "C"]])


def test_fit_auto_e_predict(seg):
    seg.fit_auto(max_depth=2, verbose=False)
    assert sum(s["is_leaf"] for s in seg.segments.values()) >= 2
    pred = seg.predict(make_df(seg.task_type, n=500, seed=9))
    assert {"segmento", "nota", "valor_regua"}.issubset(pred.columns)
    # nomes antigos task-específicos não existem mais
    assert "nota_pd" not in pred.columns and "nota_lgd" not in pred.columns


def test_faltantes_viram_bin_propria(task):
    seg = _mk(task, com_na=True, n=4000, seed=1)
    seg.grow("score", splits=[0.8])
    tem_na = any(v["is_leaf"] and v["conditions"][-1]["kind"] == "na"
                 for v in seg.segments.values())
    assert tem_na
    assert _cobertura_total(seg)


def test_grow_nao_cria_split_degenerado(task):
    df = make_df(task, n=500, seed=2)
    df["const"] = 1.0
    seg = TreeSegmenter(df, target="target", task_type=task,
                        sample_col="amostra", ref_sample="DES", verbose=False)
    seg.grow("const")
    assert len(seg.leaves()) == 1            # constante não separa


# ----------------------------------------------------------------------
# Métricas — específicas por tarefa (prova da ramificação)
# ----------------------------------------------------------------------
def test_metrics_classificacao():
    seg = _mk("classification", com_oot=True, n=6000, seed=3)
    seg.fit_auto(max_depth=3, verbose=False)
    m = seg.metrics()
    assert {"amostra", "taxa_default", "KS", "AUC", "Gini", "Acuracia", "F1"}.issubset(m.columns)
    row = m[m["amostra"] == "DES"].iloc[0]
    assert row["AUC"] > 0.6                    # régua discrimina
    assert np.isclose(row["Gini"], 2 * row["AUC"] - 1, atol=1e-6)


def test_metrics_regressao():
    seg = _mk("regression", com_oot=True, n=6000, seed=3)
    seg.fit_auto(max_depth=3, verbose=False)
    m = seg.metrics()
    assert {"amostra", "MAE", "RMSE", "R2"}.issubset(m.columns)
    row = m[m["amostra"] == "DES"].iloc[0]
    assert 0.0 <= row["MAE"] <= 1.0 and row["R2"] <= 1.0
    assert "KS" not in m.columns               # sem métricas de classificação


# ----------------------------------------------------------------------
# IV — escala específica por tarefa
# ----------------------------------------------------------------------
def test_variable_iv(seg):
    iv = seg.variable_iv("root")
    assert {"variavel", "iv", "forca", "n_bins"}.issubset(iv.columns)
    assert "score" in set(iv["variavel"])
    assert iv["iv"].max() > 0                  # alguma variável informativa


def test_variable_iv_forca_escala_por_task():
    iv_clf = _mk("classification", n=5000, seed=4).variable_iv("root")
    iv_reg = _mk("regression", n=5000, seed=4).variable_iv("root")
    # IV contínuo é numericamente bem menor que o WoE binário p/ o mesmo desenho
    assert iv_clf["iv"].max() > iv_reg["iv"].max()


def test_csi_requer_sample_col(task):
    df = make_df(task, n=500).drop(columns=["amostra"])
    seg = TreeSegmenter(df, target="target", task_type=task, verbose=False)
    seg.grow("score", splits=[0.8])
    with pytest.raises(Exception):
        seg.csi()


# ----------------------------------------------------------------------
# Persistência (inclui task_type) + régua
# ----------------------------------------------------------------------
def test_save_load_roundtrip(seg, tmp_path):
    seg.fit_auto(max_depth=2, verbose=False)
    p = str(tmp_path / "arvore.json")
    seg.save(p)
    novo = TreeSegmenter.load(p, seg.df)
    assert novo.task_type == seg.task_type
    a = seg.predict(seg.df)["segmento"]
    b = novo.predict(seg.df)["segmento"]
    assert (a.fillna("∅") == b.fillna("∅")).all()


def test_to_dict_preserva_task_type(seg):
    d = seg.to_dict()
    assert d["meta"]["task_type"] == seg.task_type
    assert d["schema"] == "yggdrasil.credit_risk.tree/1"


def test_regua_features(seg):
    seg.fit_auto(max_depth=2, verbose=False)
    feats = seg.regua_features()
    assert isinstance(feats, list) and len(feats) >= 1


def test_to_pyspark_compila(seg):
    seg.grow("score", splits=[0.8])
    code = seg.to_pyspark()
    compile(code, "<regua>", "exec")          # gera código válido
    assert "valor" in code and "def aplicar_regua" in code


# ----------------------------------------------------------------------
# Fallback de scoring p/ linhas órfãs (categoria não vista / faltante sem rota)
# ----------------------------------------------------------------------
def test_fallback_pior_nota_categoria_nao_vista(seg):
    seg.grow("garantia", splits=[["A"], ["B", "C"], ["D"]])
    novo = make_df(seg.task_type, n=300, seed=13)
    idx = novo.index[:30]
    novo.loc[idx, "garantia"] = "Z"                    # categoria não vista
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        sem = seg.predict(novo)                        # padrão: órfã fica nula
        com = seg.predict(novo, fallback="pior_nota")
    assert sem.loc[idx, "segmento"].isna().all()
    assert com["segmento"].notna().all() and com["nota"].notna().all()
    pior = seg.leaves().iloc[-1]                       # pior risco = maior nota
    assert (com.loc[idx, "segmento"] == pior["segmento"]).all()
    assert (com.loc[idx, "nota"] == pior["nota"]).all()
    # linhas com rota não mudam com o fallback
    ok = sem["segmento"].notna()
    assert (com.loc[ok, "segmento"] == sem.loc[ok, "segmento"]).all()


def test_fallback_nota_especifica_e_invalida(seg):
    seg.grow("garantia", splits=[["A"], ["B", "C"], ["D"]])
    novo = make_df(seg.task_type, n=200, seed=14)
    novo.loc[novo.index[:10], "garantia"] = "Z"
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        com = seg.predict(novo, fallback=1)            # nota específica
    assert (com.loc[novo.index[:10], "nota"] == 1).all()
    with pytest.raises(ValueError, match="não existe"):
        seg.predict(novo, fallback=99)                 # nota inexistente
    with pytest.raises(ValueError, match="fallback inválido"):
        seg.predict(novo, fallback="qualquer_coisa")


def test_to_sql_fallback_troca_else(seg):
    seg.grow("garantia", splits=[["A"], ["B", "C"], ["D"]])
    assert "ELSE NULL" in seg.to_sql()                 # padrão preservado
    sql = seg.to_sql(fallback="pior_nota")
    pior = seg.leaves().iloc[-1]
    assert "ELSE NULL" not in sql
    assert f"ELSE '{pior['segmento']}'" in sql         # segmento da folha fallback
    assert f"ELSE {int(pior['nota'])}" in sql          # nota da folha fallback
    assert "fallback" in sql                           # comentário no cabeçalho


def test_to_pyspark_fallback_troca_otherwise(seg):
    seg.grow("score", splits=[0.8])
    code = seg.to_pyspark(fallback="pior_nota")
    compile(code, "<regua>", "exec")
    pior = seg.leaves().iloc[-1]
    assert ".otherwise(F.lit(None))" not in code
    assert f".otherwise(F.lit({pior['segmento']!r}))" in code
    assert f".otherwise(F.lit({int(pior['nota'])}))" in code


def test_fallback_persistido_em_save_load(seg, tmp_path):
    seg.grow("score", splits=[0.8])
    seg.fallback = "pior_nota"
    assert seg.to_dict()["meta"]["fallback"] == "pior_nota"
    p = str(tmp_path / "arvore_fb.json")
    seg.save(p)
    novo = TreeSegmenter.load(p, seg.df)
    assert novo.fallback == "pior_nota"
    # parâmetro omitido → vale o persistido; None explícito volta ao padrão
    assert "ELSE NULL" not in novo.to_sql()
    assert "ELSE NULL" in novo.to_sql(fallback=None)


def test_n_orfas_conta_linhas_sem_rota(seg):
    assert seg.n_orfas() == 0                          # raiz cobre tudo
    # split categórico manual que NÃO cobre 'D' → linhas de 'D' ficam sem rota
    seg.grow("garantia", splits=[["A"], ["B", "C"]])
    esperado = int((seg.df["garantia"].astype(str) == "D").sum())
    assert seg.n_orfas() == esperado > 0


# ----------------------------------------------------------------------
# Apelidos de negócio por folha (leaf_names)
# ----------------------------------------------------------------------
def test_leaf_names_roundtrip_e_json_antigo(seg):
    seg.grow("score", splits=[0.8])
    folhas = [s for s, v in seg.segments.items() if v["is_leaf"]]
    seg.set_leaf_name(folhas[0], "  Baixo   risco  ")      # normaliza espaços
    assert seg.leaf_name(folhas[0]) == "Baixo risco"
    d = seg.to_dict()
    assert d["meta"]["leaf_names"] == {folhas[0]: "Baixo risco"}
    novo = TreeSegmenter.from_dict(d, seg.df)
    assert novo.leaf_name(folhas[0]) == "Baixo risco"       # roundtrip preserva
    # JSON ANTIGO (sem a chave leaf_names) carrega com apelidos vazios
    del d["meta"]["leaf_names"]
    antigo = TreeSegmenter.from_dict(d, seg.df)
    assert antigo.leaf_names == {}


def test_leaf_names_saidas_e_orfaos(seg):
    seg.grow("score", splits=[0.8])
    folhas = [s for s, v in seg.segments.items() if v["is_leaf"]]
    seg.set_leaf_name(folhas[0], "Fatia nomeada")
    # tabela de folhas ganha a coluna 'apelido' (só quando há apelido definido)
    lv = seg.leaves()
    assert "apelido" in lv.columns
    assert (lv.loc[lv["segmento"] == folhas[0], "apelido"] == "Fatia nomeada").all()
    # assign ganha a coluna '<col>_apelido' (NA nas folhas sem apelido)
    out = seg.assign("segmento")
    assert "segmento_apelido" in out.columns
    m = out["segmento"] == folhas[0]
    assert (out.loc[m, "segmento_apelido"] == "Fatia nomeada").all()
    assert out.loc[~m, "segmento_apelido"].isna().all()
    # SQL documenta o apelido como comentário no ramo do CASE
    assert "-- Fatia nomeada" in seg.to_sql()
    # apelidar algo que não é folha atual → erro
    with pytest.raises(ValueError, match="folha"):
        seg.set_leaf_name("root", "X")                      # root virou nó interno
    # fundir MUDA o sid → o apelido do sid antigo é descartado em silêncio
    seg.merge_leaf(folhas[0], side="right", verbose=False)
    assert seg.leaf_name(folhas[0]) is None
    assert seg.to_dict()["meta"]["leaf_names"] == {}
    assert "apelido" not in seg.leaves().columns


# ----------------------------------------------------------------------
# Poda / fusão / estabilidade
# ----------------------------------------------------------------------
def test_auto_merge_funde_irmas_indistinguiveis(task):
    # dois patamares bem separados → após cortes finos, irmãs indistinguíveis fundem
    df = make_df(task, n=8000, seed=5, com_oot=True)
    seg = TreeSegmenter(df, target="target", task_type=task,
                        sample_col="amostra", ref_sample="DES", verbose=False)
    seg.grow("score", splits=[0.6, 0.7, 0.8, 0.9, 1.0, 1.1])
    n0 = len(seg.leaves())
    seg.auto_merge()
    assert len(seg.leaves()) <= n0


def test_comparacao_irmas_so_terminais_adjacentes(task):
    # Task 1: irmãs só são comparáveis/fundíveis quando TERMINAIS e ADJACENTES
    # (mesma run); um nó intermediário que se expande quebra a adjacência.
    seg = _mk(task)
    seg.grow("score", splits=[0.7, 1.0])                 # 3 bins numéricos sob a raiz
    filhos = seg._ordered_direct_children("root")
    assert len(filhos) == 3
    l1, meio, l2 = filhos[0][0], filhos[1][0], filhos[2][0]
    # com os 3 terminais → 2 pares adjacentes (l1-meio, meio-l2)
    assert len(seg._adjacent_sibling_pairs("root")) == 2

    # expande a folha do MEIO → vira nó intermediário não-terminal
    seg.grow("garantia", splits=[["A", "B"], ["C", "D"]], only_segments={meio})
    assert seg.segments[meio]["is_leaf"] is False
    # agora l1 e l2 caem em runs distintas → NENHUM par adjacente sob a raiz
    assert seg._adjacent_sibling_pairs("root") == []
    # e o teste 'p (irmãs)' das folhas externas fica NaN (sem irmã adjacente)
    lv = seg.leaves(with_test=True)
    externas = lv[lv["segmento"].isin([l1, l2])]
    assert externas["p_vs_prox"].isna().all()


def test_prune_respeita_protect(seg):
    seg.grow("score", splits=[0.6, 0.8, 1.0])
    folhas = [s for s, v in seg.segments.items() if v["is_leaf"]]
    protect = set(folhas)
    seg.prune(min_repr=99.0, protect=protect)   # tudo violaria, mas está protegido
    assert set(s for s, v in seg.segments.items() if v["is_leaf"]) == protect


def test_psi_usa_segmentos_como_bins(task):
    seg = _mk(task, com_oot=True, n=6000, seed=6)
    seg.fit_auto(max_depth=2, verbose=False)
    psi = seg.psi()
    assert "psi" in psi.columns or "PSI" in psi.columns or len(psi) >= 1


# ----------------------------------------------------------------------
# Faltantes: bin própria + merge
# ----------------------------------------------------------------------
def test_merge_missing_numerico(task):
    seg = _mk(task, com_na=True, n=4000, seed=7)
    seg.grow("score", splits=[1.0])
    nums = [s for s, v in seg.segments.items()
            if v["is_leaf"] and v["conditions"][-1]["kind"] == "num"]
    assert any(v["is_leaf"] and v["conditions"][-1]["kind"] == "na"
               for v in seg.segments.values())
    seg.merge_missing(nums[-1])
    assert _cobertura_total(seg)


# ----------------------------------------------------------------------
# MOVE_CUT: mover corte numérico entre folhas-irmãs sem recolher o pai
# ----------------------------------------------------------------------
def _folha_num(seg, chave, valor):
    """sid da folha cuja última condição numérica tem ``chave`` (lo/hi) == valor."""
    return next(s for s, v in seg.segments.items()
                if v["is_leaf"] and v["conditions"]
                and v["conditions"][-1]["kind"] == "num"
                and v["conditions"][-1].get(chave) == valor)


def test_move_cut_irma_folha(seg):
    seg.grow("score", splits=[0.8])
    sid = _folha_num(seg, "hi", 0.8)
    seg.set_leaf_name(sid, "Fatia esquerda")     # apelido deve acompanhar a renomeação
    info = seg.movable_cut(sid)
    assert info["feature"] == "score" and info["cut"] == 0.8
    # preview: n dos dois lados bate com o recorte manual e NÃO altera o estado
    prev = seg.preview_move_cut(sid, 0.9)
    assert list(prev["n"]) == [int((seg.df["score"] <= 0.9).sum()),
                               int((seg.df["score"] > 0.9).sum())]
    assert "valor_DES" in prev.columns and _folha_num(seg, "hi", 0.8) == sid
    seg.move_cut(sid, 0.9, verbose=False)
    esq, dirr = _folha_num(seg, "hi", 0.9), _folha_num(seg, "lo", 0.9)
    assert int(seg.segments[esq]["mask"].sum()) == int((seg.df["score"] <= 0.9).sum())
    assert int(seg.segments[dirr]["mask"].sum()) == int((seg.df["score"] > 0.9).sum())
    assert seg.leaf_name(esq) == "Fatia esquerda"
    assert _cobertura_total(seg)


def test_move_cut_irma_com_subsplit_e_undo(task):
    seg = _mk(task)
    seg.grow("score", splits=[0.8])
    sid_dir = _folha_num(seg, "lo", 0.8)
    seg.grow("garantia", splits=[["A", "B"], ["C", "D"]], only_segments={sid_dir})
    snap = seg.to_dict()["segments"]             # p/ o "desfazer" ao final
    contagens = {s: int(v["mask"].sum()) for s, v in seg.segments.items()}

    seg.move_cut(_folha_num(seg, "hi", 0.8), 0.7, verbose=False)
    df = seg.df
    # o subtree INTEIRO da irmã (nó interno) carrega o novo lo=0.7 no caminho
    desc = {s: v for s, v in seg.segments.items() if v["depth"] == 2}
    assert len(desc) == 2
    for v in desc.values():
        assert any(c["kind"] == "num" and c.get("lo") == 0.7 for c in v["conditions"])
    # contagens batem com o recorte manual do df
    ab = df["garantia"].astype(str).isin(["A", "B"])
    folha_ab = next(v for v in desc.values()
                    if v["conditions"][-1]["kind"] == "cat"
                    and v["conditions"][-1]["cats"] == ["A", "B"])
    assert int(folha_ab["mask"].sum()) == int(((df["score"] > 0.7) & ab).sum())
    assert int(seg.segments[_folha_num(seg, "hi", 0.7)]["mask"].sum()) == \
        int((df["score"] <= 0.7).sum())
    assert _cobertura_total(seg)
    # "undo" (mesmo caminho do desfazer da UI): restaura estrutura e contagens
    seg._load_segments(snap)
    assert {s: int(v["mask"].sum()) for s, v in seg.segments.items()} == contagens


def test_move_cut_validacoes(seg):
    seg.grow("score", splits=[0.6, 0.8])
    meio = _folha_num(seg, "hi", 0.8)                    # folha (0.6, 0.8]
    with pytest.raises(ValueError, match="intervalo"):
        seg.move_cut(meio, 0.5)                          # abaixo do lo da folha
    with pytest.raises(ValueError, match="intervalo"):
        seg.move_cut(meio, 0.6)                          # exclusivo: == lo não vale
    with pytest.raises(ValueError, match="ESQUERDA"):
        seg.move_cut(_folha_num(seg, "lo", 0.8), 1.0)    # última faixa (hi = inf)
    with pytest.raises(ValueError, match="sem linhas"):  # atômico: lado esvaziado
        seg.move_cut(_folha_num(seg, "hi", 0.6), 0.05)   # score mínimo ≈ 0.3
    # nada mudou após as tentativas inválidas
    assert {0.6, 0.8} == {v["conditions"][-1]["hi"] for v in seg.segments.values()
                          if v["is_leaf"] and v["conditions"][-1]["kind"] == "num"
                          and np.isfinite(v["conditions"][-1]["hi"])}


def test_move_cut_exige_split_numerico(seg):
    seg.grow("garantia", splits=[["A"], ["B", "C", "D"]])
    folha_cat = next(s for s, v in seg.segments.items()
                     if v["is_leaf"] and v["conditions"][-1]["kind"] == "cat")
    with pytest.raises(ValueError, match="num"):
        seg.move_cut(folha_cat, 0.5)
    assert seg.movable_cut(folha_cat) is None


# ----------------------------------------------------------------------
# Backtest / calibração / monotonicidade
# ----------------------------------------------------------------------
def test_backtest_por_safra(task):
    seg = _mk(task, com_oot=True, n=6000, seed=8)
    seg.fit_auto(max_depth=2, verbose=False)
    bt = seg.backtest("dt_ref")
    assert {"periodo", "valor_previsto", "valor_realizado", "gap", "status"}.issubset(bt.columns)


def test_calibration_table(task):
    seg = _mk(task, com_oot=True, n=6000, seed=8)
    seg.fit_auto(max_depth=2, verbose=False)
    ct = seg.calibration_table()
    assert {"valor_previsto", "valor_realizado", "gap"}.issubset(ct.columns)


def test_monotonicity_report(task):
    seg = _mk(task, com_oot=True, n=6000, seed=8)
    seg.fit_auto(max_depth=3, verbose=False)
    mr = seg.monotonicity_report()
    assert {"amostra", "monotonico", "n_inversoes"}.issubset(mr.columns)


# ----------------------------------------------------------------------
# Plots — comuns e específicos por tarefa (gated por matplotlib)
# ----------------------------------------------------------------------
def test_plot_tree_gera_imagem(seg):
    pytest.importorskip("matplotlib")
    seg.fit_auto(max_depth=2, verbose=False)
    fig = seg.plot_tree()
    assert fig is not None


def test_plot_tree_hitmap(seg):
    """PNG + caixa clicável por nó (origem no canto superior esquerdo), base do
    preview interativo da UI: todo segmento tem caixa, dentro da imagem, com o
    flag de folha correto; folhas não se sobrepõem no eixo x (layout X_GAP)."""
    pytest.importorskip("matplotlib")
    seg.fit_auto(max_depth=2, verbose=False)
    out = seg.plot_tree_hitmap(dpi=96)
    assert out["png"][:8] == b"\x89PNG\r\n\x1a\n"      # PNG válido
    assert out["width"] > 0 and out["height"] > 0
    assert set(out["nodes"]) == set(seg.segments)      # 1 caixa por segmento
    for sid, b in out["nodes"].items():
        assert 0 <= b["x0"] < b["x1"] <= out["width"]
        assert 0 <= b["y0"] < b["y1"] <= out["height"]
        assert b["is_leaf"] == seg.segments[sid]["is_leaf"]
    folhas = sorted((b for b in out["nodes"].values() if b["is_leaf"]),
                    key=lambda b: b["x0"])
    assert len(folhas) >= 2
    for a, b in zip(folhas, folhas[1:]):
        assert a["x1"] <= b["x0"]                      # sem sobreposição horizontal


def test_plots_especificos_classificacao():
    pytest.importorskip("matplotlib")
    seg = _mk("classification", com_oot=True, n=4000, seed=3)
    seg.fit_auto(max_depth=2, verbose=False)
    assert seg.plot_roc() is not None
    assert seg.plot_ks() is not None
    assert seg.plot_score_distribution() is not None


def test_plots_especificos_regressao():
    pytest.importorskip("matplotlib")
    seg = _mk("regression", com_oot=True, n=4000, seed=3)
    seg.fit_auto(max_depth=2, verbose=False)
    assert seg.plot_leaf_boxplots() is not None
    assert seg.plot_target_hist() is not None
    assert seg.plot_leaf_value_hist() is not None


def test_plot_feature_value(seg):
    pytest.importorskip("matplotlib")
    fig = seg.plot_feature_value("score")
    assert fig is not None


def test_plot_cap_lift_metricas_por_safra(task):
    """Smoke dos três diagnósticos novos nos DOIS task_type (paridade clf×reg):
    CAP com AR na legenda, lift por decil e métricas por safra (clf: KS/AUC ·
    reg: R²/RMSE), sempre com o alvo previsto da folha como score."""
    pytest.importorskip("matplotlib")
    seg = _mk(task, com_oot=True, n=4000, seed=3)
    seg.fit_auto(max_depth=2, verbose=False)
    fig = seg.plot_cap()
    labels = [t.get_text() for t in fig.axes[0].get_legend().get_texts()]
    assert any("AR=" in t for t in labels)             # curva útil por amostra
    assert seg.plot_lift() is not None
    ms = seg.metrics_by_safra(time_col="dt_ref")
    want = ({"safra", "n", "taxa_evento", "auc", "ks", "gini"}
            if task == "classification"
            else {"safra", "n", "previsto_medio", "realizado_medio", "mae", "rmse", "r2"})
    assert want.issubset(ms.columns) and len(ms) >= 2
    assert seg.plot_metrics_by_safra(time_col="dt_ref") is not None


def test_metrics_by_safra_sem_time_col_erro(task):
    seg = _mk(task)                                    # sem date_col configurado
    with pytest.raises(ValueError, match="time_col"):
        seg.metrics_by_safra()


# ----------------------------------------------------------------------
# PySpark (gated)
# ----------------------------------------------------------------------
def test_log_to_mlflow_metrricas_por_task(task, tmp_path):
    pytest.importorskip("mlflow")
    import mlflow
    mlflow.set_tracking_uri((tmp_path / "mlruns").as_uri())
    seg = _mk(task, com_oot=True, n=4000, seed=3)
    seg.fit_auto(max_depth=3, verbose=False)
    rid = seg.log_to_mlflow(experiment=f"t_{task}", verbose=False)
    run = mlflow.get_run(rid)
    # variáveis + profundidade nos params; n_variaveis nas métricas
    assert "variaveis" in run.data.params and "profundidade" in run.data.params
    assert "n_variaveis" in run.data.metrics
    # PSI por amostra (OOT) + métricas conforme o task
    assert any(k.startswith("psi_") for k in run.data.metrics)
    want = ["ks_DES", "auc_DES", "gini_DES"] if task == "classification" else ["mae_DES", "rmse_DES", "r2_DES"]
    assert all(k in run.data.metrics for k in want)


def test_to_sql_case_when(seg):
    seg.fit_auto(max_depth=2, verbose=False)
    sql = seg.to_sql(table="carteira")
    assert "CASE" in sql and "carteira" in sql
    assert "AS segmento" in sql and "AS folha" in sql and "AS valor_previsto" in sql
    # uma cláusula WHEN por folha
    assert sql.count("WHEN") >= 2 * len(seg.leaves())


def test_feature_importance(seg):
    seg.fit_auto(max_depth=3, verbose=False)
    fi = seg.feature_importance()
    assert {"variavel", "n_splits", "importancia"}.issubset(fi.columns)
    # só lista variáveis que entraram na árvore
    assert len(fi) >= 1 and (fi["n_splits"] >= 1).all()
    if "importancia_%" in fi.columns and fi["importancia_%"].sum() > 0:
        assert abs(fi["importancia_%"].sum() - 100.0) < 1.0


def test_plot_importance_bar_estilo(task):
    """O gráfico de importância relativa: sem eixo x (só os rótulos por barra) e
    degradê por magnitude — mais importante steelblue, menos importante crimson."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.colors as mc
    seg = _mk(task, n=6000, seed=3)
    seg.fit_auto(max_depth=3, verbose=False)
    fig = seg.plot_importance_bar()
    ax = fig.axes[0]
    assert list(ax.get_xticks()) == []                 # sem labels no eixo x
    fi = seg.feature_importance()
    if len(fi) >= 2:                                    # com ≥2 variáveis há degradê
        vals = fi["importancia_%"].to_numpy()
        cores = [p.get_facecolor()[:3] for p in ax.patches]
        sb, cr = mc.to_rgb("steelblue"), mc.to_rgb("crimson")
        assert tuple(round(x, 2) for x in cores[int(vals.argmax())]) == \
            tuple(round(x, 2) for x in sb)             # maior → steelblue
        assert tuple(round(x, 2) for x in cores[int(vals.argmin())]) == \
            tuple(round(x, 2) for x in cr)             # menor → crimson
    import matplotlib.pyplot as plt
    plt.close(fig)


def test_suggest_splits(task):
    seg = _mk(task, com_oot=True, n=6000, seed=3)
    sug = seg.suggest_splits(top=3)
    assert {"variavel", "n_bins", "iv", "passa_teste", "p_valor"}.issubset(sug.columns)
    assert any(c.startswith("psi_") for c in sug.columns)     # PSI por amostra (OOT)
    assert len(sug) >= 1 and sug["passa_teste"].dtype == bool


CRITERIOS = {
    "classification": ["gini", "entropy", "ks", "iv", "chi2"],
    "regression": ["variance", "mae", "ftest"],
}


def test_fit_auto_por_criterio(task):
    for crit in CRITERIOS[task]:
        seg = _mk(task, n=5000, seed=4)
        seg.fit_auto(max_depth=3, criterion=crit, verbose=False)
        nf = sum(s["is_leaf"] for s in seg.segments.values())
        assert nf >= 2, f"critério {crit} não dividiu"
        assert _cobertura_total(seg)


def test_grow_por_criterio_e_binario(seg):
    # critério != optbin faz split BINÁRIO (2 filhos) por folha numérica
    seg.grow("score", criterion="gini" if seg.task_type == "classification" else "variance")
    assert len(seg.leaves()) == 2


def test_diff_trees(task):
    a = _mk(task, n=6000, seed=5)
    a.fit_auto(max_depth=3, verbose=False)
    b = _mk(task, df=a.df.copy())
    b.fit_auto(max_depth=1, verbose=False)
    d = a.diff_trees(b)
    assert 0.0 <= d["concordancia"] <= 1.0
    assert d["migracao"].shape[0] >= 1 and d["migracao"].shape[1] >= 1
    assert {"métrica", "árvore A", "árvore B"}.issubset(d["resumo"].columns)


def test_diff_trees_task_incompativel():
    a = _mk("classification", n=1000)
    b = _mk("regression", df=make_df("regression", n=1000))
    with pytest.raises(ValueError, match="task_type"):
        a.diff_trees(b)


def test_from_dict_mask_cache_compartilhado(task):
    """`from_dict(..., mask_cache=...)` (cenários em memória): reconstrução sobre
    o MESMO df reusa as máscaras vivas registradas por `_prime_mask_cache` —
    mesmo objeto por segmento (nada recomputado) e cache compartilhado."""
    seg = _mk(task)
    seg.grow("score", splits=[0.8])
    d = seg.to_dict()
    cache = seg._prime_mask_cache()               # devolve o cache primado
    novo = TreeSegmenter.from_dict(d, seg.df, mask_cache=cache)
    assert set(novo.segments) == set(seg.segments)
    for sid, s in seg.segments.items():           # máscara REUSADA (identidade)
        assert novo.segments[sid]["mask"] is s["mask"]
    assert novo._mask_cache_by_conds is seg._mask_cache_by_conds
    # e o diff das duas árvores idênticas concorda 100%
    assert seg.diff_trees(novo)["concordancia"] == pytest.approx(1.0)


# ----------------------------------------------------------------------
# Excel multi-abas (gated por openpyxl — dependência OPCIONAL)
# ----------------------------------------------------------------------
def test_to_excel_multiabas_roundtrip(task, tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.utils import get_column_letter
    seg = _mk(task, com_oot=True, n=4000, seed=3)
    seg.fit_auto(max_depth=2, verbose=False)
    p = str(tmp_path / "arvore.xlsx")
    assert seg.to_excel(p, table="carteira") == p
    wb = openpyxl.load_workbook(p)
    for aba in ["Folhas", "Métricas por amostra", "PSI", "IV por variável",
                "Calibração", "Régua SQL"]:
        assert aba in wb.sheetnames, f"aba ausente: {aba}"
    assert wb["Folhas"].freeze_panes == "A2"          # cabeçalho congelado
    lv = pd.read_excel(p, sheet_name="Folhas")
    assert len(lv) == len(seg.leaves())
    # representatividade exportada como FRAÇÃO (0–1) com formato de % no Excel
    rep = [c for c in lv.columns if str(c).startswith("repr_")]
    assert rep and lv[rep].fillna(0).le(1.0).all().all()
    j = [c.value for c in wb["Folhas"][1]].index(rep[0]) + 1
    assert wb["Folhas"][f"{get_column_letter(j)}2"].number_format == "0.00%"
    # aba PSI traz o resumo E o detalhe por folha logo abaixo
    textos = [str(c.value) for row in wb["PSI"].iter_rows() for c in row
              if c.value is not None]
    assert any("Detalhe" in t for t in textos)
    # régua SQL como texto (uma linha por célula da coluna A)
    sql = "\n".join(str(c.value) for c in wb["Régua SQL"]["A"] if c.value is not None)
    assert "CASE" in sql and "carteira" in sql


def test_to_excel_sem_sample_col_pula_psi(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    df = make_df("classification", n=800).drop(columns=["amostra"])
    seg = TreeSegmenter(df, target="target", task_type="classification", verbose=False)
    seg.grow("score", splits=[0.8])
    p = str(tmp_path / "sem_amostra.xlsx")
    seg.to_excel(p)
    wb = openpyxl.load_workbook(p)
    assert "PSI" not in wb.sheetnames and "Folhas" in wb.sheetnames


def test_to_excel_sem_openpyxl_erro_amigavel(tmp_path, monkeypatch):
    """Sem o openpyxl (opcional), o to_excel sobe ImportError com a instrução
    de instalação — nunca um traceback críptico de dentro do pandas."""
    import builtins
    real_import = builtins.__import__

    def sem_openpyxl(name, *a, **kw):
        if name == "openpyxl" or name.startswith("openpyxl."):
            raise ImportError("No module named 'openpyxl'")
        return real_import(name, *a, **kw)

    monkeypatch.setattr(builtins, "__import__", sem_openpyxl)
    seg = _mk("classification", n=400)
    with pytest.raises(ImportError, match="pip install openpyxl"):
        seg.to_excel(str(tmp_path / "x.xlsx"))


def test_apply_spark_roundtrip(task):
    pytest.importorskip("pyspark")
    from pyspark.sql import SparkSession
    spark = SparkSession.builder.master("local[1]").appName("t").getOrCreate()
    try:
        seg = _mk(task, n=3000, seed=9)
        seg.grow("score", splits=[0.8])
        sdf = spark.createDataFrame(seg.df[["score", "garantia"]])
        out = seg.apply_spark(sdf).toPandas()
        assert {"segmento", "nota", "valor_regua"}.issubset(out.columns)
    finally:
        spark.stop()


def test_apply_table_pandas_progresso(seg):
    """`apply_table` no caminho pandas puro (sem Spark), de ponta a ponta: aplica
    a régua num DataFrame em memória, devolve ``(saida, resumo)`` com o resumo
    por folha calculado no mesmo passo e o ``progress_callback`` registra as
    etapas — e, mesmo QUEBRANDO, não derruba a aplicação (progresso é
    cosmético)."""
    seg.grow("score", splits=[0.8])
    eventos = []

    def cb(key, label, status, detail=""):
        eventos.append((key, status))
        raise RuntimeError("callback quebrado não pode derrubar a aplicação")

    novo = make_df(seg.task_type, n=400, seed=21)
    out, resumo = seg.apply_table(novo, progress_callback=cb)
    # saída pandas: a base original + as colunas da régua, mesmo nº de linhas
    assert {"segmento", "nota", "valor_regua"}.issubset(out.columns)
    assert set(novo.columns).issubset(out.columns) and len(out) == len(novo)
    assert out["nota"].notna().all()               # cobertura total neste desenho
    # resumo por folha: contagem cobre todas as linhas e a fração soma 1
    assert list(resumo.columns) == ["nota", "linhas", "pct"]
    assert int(resumo["linhas"].sum()) == len(novo)
    assert float(resumo["pct"].sum()) == pytest.approx(1.0)
    # etapas na ordem run→ok, sem as etapas Spark (ler tabela / gravar)
    assert eventos == [("aplicar", "run"), ("aplicar", "ok"),
                       ("resumo", "run"), ("resumo", "ok"), ("done", "ok")]


# ----------------------------------------------------------------------
# PESOS DE EXPOSIÇÃO (weight_col) — visão dupla contratos × saldo (fase 1)
# ----------------------------------------------------------------------
def _df_pesos(task, n=1500, seed=7):
    """Base com peso (saldo) CONCENTRADO nas linhas de score alto (⇒ o % do saldo
    da folha diverge do % de contratos) e também nas garantias piores (⇒ dentro da
    MESMA folha o alvo ponderado difere do alvo simples)."""
    df = make_df(task, n=n, seed=seed, com_oot=True)
    df["saldo"] = (np.where(df["score"] > 0.9, 900.0, 100.0)
                   * np.where(df["garantia"].isin(["C", "D"]), 3.0, 1.0))
    return df


def _seg_pesos(task, df=None, **kw):
    seg = _mk(task, df=_df_pesos(task) if df is None else df,
              weight_col="saldo", **kw)
    seg.grow("score", splits=[0.9])
    return seg


def test_weight_col_saldo_diverge_de_contratos_e_bate_np_average(task):
    """Com pesos concentrados, `saldo_%` diverge de `repr_%` na mesma folha, e o
    alvo ponderado reproduz exatamente o np.average calculado à mão."""
    seg = _seg_pesos(task)
    lv = seg.leaves()
    assert {"saldo_%", "valor_medio_pond", "valor_pond_DES",
            "valor_pond_OOT"}.issubset(lv.columns)
    # as duas visões fecham 100%, mas folha a folha são bem diferentes
    assert lv["saldo_%"].sum() == pytest.approx(100.0, abs=0.3)
    assert (lv["saldo_%"] - lv["repr_%"]).abs().max() > 5.0
    for sid, esperado_pond in zip(lv["segmento"], lv["valor_medio_pond"]):
        m = seg.segments[sid]["mask"] & (seg.df["amostra"] == "DES")
        y = seg.df.loc[m, "target"].to_numpy(dtype="float64")
        w = seg.df.loc[m, "saldo"].to_numpy(dtype="float64")
        assert esperado_pond == pytest.approx(np.average(y, weights=w), abs=1e-4)
        # weight_share = fração do saldo total nesta folha
        share = 100 * seg.df.loc[seg.segments[sid]["mask"], "saldo"].sum() \
            / seg.df["saldo"].sum()
        assert seg.weight_share(sid) == pytest.approx(share, abs=1e-9)
    # ponderar muda o número (pesos não uniformes dentro da folha)
    assert (lv["valor_medio"] - lv["valor_medio_pond"]).abs().max() > 1e-3


def test_weight_col_fora_da_modelagem_e_split_nao_ponderado(task):
    """FASE 1: o peso NÃO é variável candidata e NÃO entra no critério de split —
    a árvore sai idêntica à construída sem `weight_col`."""
    df = _df_pesos(task)
    com = _mk(task, df=df, weight_col="saldo")
    sem = _mk(task, df=df.drop(columns=["saldo"]))
    assert "saldo" not in set(com.variable_iv("root")["variavel"])
    com.fit_auto(max_depth=2, verbose=False)
    sem.fit_auto(max_depth=2, verbose=False)
    assert list(com.leaves()["descricao"]) == list(sem.leaves()["descricao"])


def test_sem_weight_col_nada_muda(seg):
    """Sem `weight_col`, as colunas ponderadas não existem e os helpers devolvem
    NaN — o caso padrão fica intocado."""
    seg.grow("score", splits=[0.9])
    lv = seg.leaves()
    assert seg.weight_col is None
    assert not [c for c in lv.columns if c == "saldo_%" or "pond" in str(c)]
    sid = lv["segmento"].iloc[0]
    assert np.isnan(seg.weight_share(sid)) and np.isnan(seg.weighted_value(sid))


def test_weight_col_nan_fica_fora_das_somas(task):
    """Peso ausente (NaN) não entra nem na soma do saldo nem na média ponderada."""
    df = _df_pesos(task, n=900, seed=11)
    df.loc[df.index[:250], "saldo"] = np.nan
    seg = _seg_pesos(task, df=df)
    lv = seg.leaves()
    assert lv["saldo_%"].sum() == pytest.approx(100.0, abs=0.3)
    sid = lv["segmento"].iloc[0]
    m = (seg.segments[sid]["mask"] & (seg.df["amostra"] == "DES")
         & seg.df["saldo"].notna())
    esperado = np.average(seg.df.loc[m, "target"], weights=seg.df.loc[m, "saldo"])
    assert seg.weighted_value(sid, "DES") == pytest.approx(esperado, abs=1e-9)


def test_weight_col_invalida_levanta_erro(task):
    df = _df_pesos(task, n=400)
    with pytest.raises(ValueError, match="não está no DataFrame"):
        _mk(task, df=df, weight_col="nao_existe")
    txt = df.assign(saldo_txt="x")
    with pytest.raises(ValueError, match="numérica"):
        _mk(task, df=txt, weight_col="saldo_txt")
    neg = df.copy()
    neg.loc[neg.index[:5], "saldo"] = -1.0
    with pytest.raises(ValueError, match="NEGATIVO"):
        _mk(task, df=neg, weight_col="saldo")


def test_weight_col_roundtrip_to_dict_e_json_antigo(task):
    seg = _seg_pesos(task)
    d = seg.to_dict()
    assert d["meta"]["weight_col"] == "saldo"
    novo = TreeSegmenter.from_dict(d, seg.df)
    assert novo.weight_col == "saldo" and "saldo_%" in novo.leaves().columns
    # JSON ANTIGO (sem a chave) carrega com weight_col=None
    antigo_meta = dict(d)
    antigo_meta["meta"] = {k: v for k, v in d["meta"].items() if k != "weight_col"}
    antigo = TreeSegmenter.from_dict(antigo_meta, seg.df)
    assert antigo.weight_col is None and "saldo_%" not in antigo.leaves().columns
    # df sem a coluna de peso: degrada p/ None em vez de quebrar a carga
    degradado = TreeSegmenter.from_dict(d, seg.df.drop(columns=["saldo"]))
    assert degradado.weight_col is None


def test_to_excel_traz_visao_dupla(task, tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    seg = _seg_pesos(task)
    p = str(tmp_path / "pesos.xlsx")
    seg.to_excel(p)
    ws = openpyxl.load_workbook(p)["Folhas"]
    cab = [c.value for c in ws[1]]
    assert "saldo_%" in cab and "valor_medio_pond" in cab
    col = cab.index("saldo_%") + 1
    # % do saldo sai como FRAÇÃO 0–1 com formato de porcentagem no Excel
    assert 0.0 <= ws.cell(row=2, column=col).value <= 1.0
    assert ws.cell(row=2, column=col).number_format == "0.00%"
